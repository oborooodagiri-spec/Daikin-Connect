import os
import sys
import json
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from PIL import Image

def get_val(sheet, row, col):
    val = sheet.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()

def convert_excel_to_json(excel_path, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    result = []
    
    # Process "AHU & AC SPLIT"
    if "AHU & AC SPLIT" in wb.sheetnames:
        sheet = wb["AHU & AC SPLIT"]
        image_loader = SheetImageLoader(sheet)
        
        for row in range(6, sheet.max_row + 1):
            if get_val(sheet, row, 4): # 'TANGGAL'
                item = {
                    "tab": "AHU & AC SPLIT",
                    "no": get_val(sheet, row, 3),
                    "tanggal": get_val(sheet, row, 4),
                    "lantai": get_val(sheet, row, 5),
                    "tenant_area": get_val(sheet, row, 6),
                    "kode_unit": get_val(sheet, row, 7),
                    "brand": get_val(sheet, row, 8),
                    "model": get_val(sheet, row, 9),
                    "corrective_action": get_val(sheet, row, 10),
                    "status": get_val(sheet, row, 17),
                    "photos": []
                }
                
                for col_idx, col_letter in enumerate(['K', 'L', 'M', 'N', 'O', 'P']):
                    cell_coord = f"{col_letter}{row}"
                    try:
                        if image_loader.image_in(cell_coord):
                            img = image_loader.get(cell_coord)
                            if img.mode in ('RGBA', 'P'):
                                img = img.convert('RGB')
                            filename = f"ahu_{row}_{col_letter}.jpg"
                            filepath = os.path.join(output_dir, filename)
                            img.save(filepath, "JPEG")
                            item["photos"].append(filename)
                    except Exception as e:
                        pass
                
                if item["tenant_area"]:
                    result.append(item)
                    
    # Process "FCU"
    if "FCU" in wb.sheetnames:
        sheet = wb["FCU"]
        image_loader = SheetImageLoader(sheet)
        
        for row in range(3, sheet.max_row + 1):
            if get_val(sheet, row, 3): # 'Tanggal'
                item = {
                    "tab": "FCU",
                    "no": get_val(sheet, row, 2),
                    "tanggal": get_val(sheet, row, 3),
                    "lantai": get_val(sheet, row, 4),
                    "tenant_area": get_val(sheet, row, 5),
                    "kode_unit": get_val(sheet, row, 6),
                    "brand": get_val(sheet, row, 7),
                    "model": get_val(sheet, row, 8),
                    "remarks": get_val(sheet, row, 9),
                    "qty": get_val(sheet, row, 10),
                    "corrective_action": get_val(sheet, row, 11),
                    "status": get_val(sheet, row, 18),
                    "photos": []
                }
                
                for col_idx, col_letter in enumerate(['L', 'M', 'N', 'O', 'P', 'Q']):
                    cell_coord = f"{col_letter}{row}"
                    try:
                        if image_loader.image_in(cell_coord):
                            img = image_loader.get(cell_coord)
                            if img.mode in ('RGBA', 'P'):
                                img = img.convert('RGB')
                            filename = f"fcu_{row}_{col_letter}.jpg"
                            filepath = os.path.join(output_dir, filename)
                            img.save(filepath, "JPEG")
                            item["photos"].append(filename)
                    except Exception as e:
                        pass
                
                if item["tenant_area"]:
                    result.append(item)

    with open(os.path.join(output_dir, 'extracted_data.json'), 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
        
    print(f"Extracted {len(result)} items to {output_dir}/extracted_data.json")

if __name__ == "__main__":
    excel_path = sys.argv[1]
    output_dir = sys.argv[2]
    convert_excel_to_json(excel_path, output_dir)
